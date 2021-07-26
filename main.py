import requests
import json
import pandas as pd
from openpyxl import Workbook
import os
import time


asset_contract_address = "0x3fe1a4c1481c8351e91b64d5c398b159de07cbc5"
token_idList = [
    2560,
  794,
 8760,
 8048
]


def getTokenStuff(final_dictionary):
    traits = final_dictionary['traits']
    myTraits = {}
    mylist = []
    for type in traits:
        for key, value in type.items():
            if key == 'trait_type':
                mylist = [val for val in type.values() if val !=
                          value and val != None]
                myTraits[value] = mylist
    normalizedTraits = pd.json_normalize(myTraits)
    return normalizedTraits


def getSummaryStuff(token):
    output = pd.DataFrame()
    fetchSingleAsset = 'https://api.opensea.io/api/v1/asset/{}/{}'.format(
        asset_contract_address, token)
    response = requests.request("GET", fetchSingleAsset)
    final_dictionary = json.loads(response.text)

    name = final_dictionary['asset_contract']['name']
    total_supply = final_dictionary['collection']['stats']['total_supply']
    total_owners = final_dictionary['collection']['stats']['num_owners']
    floor_price = final_dictionary['collection']['stats']['floor_price']
    total_volume_traded = final_dictionary['collection']['stats']['total_volume']
    description = final_dictionary['asset_contract']['description']
    summaryInfo = {
        "Name": name,
        "Total Items": total_supply,
        "Total Owners": total_owners,
        "Floor Price": floor_price,
        "Total Volume Traded": total_volume_traded,
        "Description": description,
    }
    output = output.append(summaryInfo, ignore_index=True)
    return output


def calculateETHprice(final_dictionary):
    price = 0
    eth_price_calc = 1000000000000000000
    if len(final_dictionary['orders']) == 0:
        final_dictionary['orders'] = "None"
    
    if len(final_dictionary['owner']) == 0:
        final_dictionary['owner'] = "None"
    else:
        last_current_price = len(final_dictionary['orders']) - 1
        for key,value in final_dictionary['orders'][last_current_price].items():
            if key == 'current_price':
                price = float(value) / eth_price_calc
    return price

def getTheMainStuff():
    output = pd.DataFrame()
    for token in token_idList:
        print(token)
        fetchSingleAsset = 'https://api.opensea.io/api/v1/asset/{}/{}'.format(
            asset_contract_address, token)

        time.sleep(1)
        response = requests.request("GET", fetchSingleAsset)
        final_dictionary = json.loads(response.text)

        # doing this in a hacky way right now,
        makeBool = True
        for key, value in final_dictionary.items():
            if key == 'last_sale' and value is None:
                makeBool = False
                break
        
        for key, value in final_dictionary['owner'].items():
            if key == 'user' and value is None:
                final_dictionary['owner']['user'] = 'None'
                makeBool = False

            # print(type(final_dictionary['owner']['user']))

        ownerBool = False
        if type(final_dictionary['owner']['user']) == dict:
            if final_dictionary['owner']['user']['username'] is None:
                final_dictionary['owner']['user'] = 'None'
        else:
            ownerBool = True
           

        current_price = calculateETHprice(final_dictionary)

        owner_user = final_dictionary['owner']['user']



        if makeBool:
            print('1ST')
            generalInfo = {
                'Collection Name': final_dictionary['collection']['primary_asset_contracts'][0]['name'],
                'Project Contract Address': final_dictionary['collection']['primary_asset_contracts'][0]['address'],
                'External Link ': final_dictionary['collection']['primary_asset_contracts'][0]['external_link'],
                'Created Date': final_dictionary['collection']['primary_asset_contracts'][0]['created_date'],
                'Schema Name': final_dictionary['collection']['primary_asset_contracts'][0]['schema_name'],
                'Symbol': final_dictionary['collection']['primary_asset_contracts'][0]['symbol'],
                'Creator User': final_dictionary['creator']['user']['username'],
                'NFT Name': final_dictionary['name'],
                'Token ID': final_dictionary['token_id'],
                'Auction Type': final_dictionary['last_sale']['auction_type'],
                'Total Price': final_dictionary['last_sale']['total_price'],
                'Last Sale Creation Date': final_dictionary['last_sale']['created_date'],
                'Quantity': final_dictionary['last_sale']['quantity'],
                'Telegram URL': final_dictionary['collection']['telegram_url'],
                'Twitter User': final_dictionary['collection']['twitter_username'],
                'Instagram User': final_dictionary['collection']['instagram_username'],
                'Discord URL': final_dictionary['collection']['discord_url'],
                # ETH Price
                'Current Price': current_price,
                'Total NFTs in Collection': final_dictionary['collection']['stats']['total_supply'],
                'Owner': owner_user['username'],
                'Permalink': final_dictionary['permalink']

            }
        elif makeBool == False:
            print('2ND')
            generalInfo = {
                'Collection Name': final_dictionary['collection']['primary_asset_contracts'][0]['name'],
                'Project Contract Address': final_dictionary['collection']['primary_asset_contracts'][0]['address'],
                'External Link ': final_dictionary['collection']['primary_asset_contracts'][0]['external_link'],
                'Created Date': final_dictionary['collection']['primary_asset_contracts'][0]['created_date'],
                'Schema Name': final_dictionary['collection']['primary_asset_contracts'][0]['schema_name'],
                'Symbol': final_dictionary['collection']['primary_asset_contracts'][0]['symbol'],
                'Creator User': final_dictionary['creator']['user']['username'],
                'NFT Name': final_dictionary['name'],
                'Token ID': final_dictionary['token_id'],
                'Auction Type': final_dictionary['last_sale'],
                'Total Price': final_dictionary['last_sale'],
                'Last Sale Creation Date': final_dictionary['last_sale'],
                'Quantity': final_dictionary['last_sale'],
                'Telegram URL': final_dictionary['collection']['telegram_url'],
                'Twitter User': final_dictionary['collection']['twitter_username'],
                'Instagram User': final_dictionary['collection']['instagram_username'],
                'Discord URL': final_dictionary['collection']['discord_url'],
                'Current Price': current_price,  # ETH Price
                'Total NFTs in Collection': final_dictionary['collection']['stats']['total_supply'],
                'Owner': final_dictionary['owner']['user'],
                'Permalink': final_dictionary['permalink']

            }
        if ownerBool:
            print('3RD')
            generalInfo = {'Collection Name': final_dictionary['collection']['primary_asset_contracts'][0]['name'],
                'Project Contract Address': final_dictionary['collection']['primary_asset_contracts'][0]['address'],
                'External Link ': final_dictionary['collection']['primary_asset_contracts'][0]['external_link'],
                'Created Date': final_dictionary['collection']['primary_asset_contracts'][0]['created_date'],
                'Schema Name': final_dictionary['collection']['primary_asset_contracts'][0]['schema_name'],
                'Symbol': final_dictionary['collection']['primary_asset_contracts'][0]['symbol'],
                'Creator User': final_dictionary['creator']['user']['username'],
                'NFT Name': final_dictionary['name'],
                'Token ID': final_dictionary['token_id'],
                'Auction Type': final_dictionary['last_sale']['auction_type'],
                'Total Price': final_dictionary['last_sale']['total_price'],
                'Last Sale Creation Date': final_dictionary['last_sale']['created_date'],
                'Quantity': final_dictionary['last_sale']['quantity'],
                'Telegram URL': final_dictionary['collection']['telegram_url'],
                'Twitter User': final_dictionary['collection']['twitter_username'],
                'Instagram User': final_dictionary['collection']['instagram_username'],
                'Discord URL': final_dictionary['collection']['discord_url'],
                # ETH Price
                'Current Price': current_price,
                'Total NFTs in Collection': final_dictionary['collection']['stats']['total_supply'],
                'Owner': final_dictionary['owner']['user'],
                'Permalink': final_dictionary['permalink']}


        mytraits = getTokenStuff(final_dictionary)
        for trait_type, valAndCountList in mytraits.items():
            generalInfo['Traits: ' + trait_type] = valAndCountList[0][0]
            generalInfo['Traits: ' + trait_type + ' (#) '] = valAndCountList[0][1]
            generalInfo['Traits ' + trait_type + ' %']  = valAndCountList[0][1] / final_dictionary['collection']['stats']['total_supply']

        output = output.append(generalInfo, ignore_index=True)
    return output


writer = pd.ExcelWriter('Final.xlsx')
df = getTheMainStuff()
summaryDf = getSummaryStuff(5477)
df.to_excel(writer, sheet_name='General Info', index=True)
summaryDf.to_excel(writer, sheet_name='Summary', index=True)

writer.save()


# def getAllDataFromAssetAndConverToExcelSheets():
#     wb = Workbook()
#     defaultSheet = wb.active
#     defaultSheet.title = "assetData"
#     mylist = ['asset_contract', 'collection', 'traits', 'owner',
#             'creator', 'orders', 'top_ownerships', 'last_sale']

#     for sections in asset:
#         if sections in mylist and isinstance(asset[sections], dict):
#             sec = wb.create_sheet(sections)
#             for key, value in asset[sections].items():
#                 print(key, '-->', value, "0")
#                 if isinstance(value, list):
#                     innerSec = wb.create_sheet(sections+"_"+key)
#                     for item in value:
#                         for k, v in item.items():
#                             print(k, '-->', v, "1")
#                             innerSec.append([k, v])
#                 elif isinstance(value, dict):
#                     innerSec = wb.create_sheet(sections+"_"+key)
#                     for k, v in value.items():
#                         if isinstance(v, dict):
#                             for m, n in v.items():
#                                 if isinstance(n, dict):
#                                     for a, b in n.items():
#                                         print(a, '-->', b, "2")
#                                         innerSec.append([a, b])
#                                 else:
#                                     print(m, '-->', n, "3")
#                                     innerSec.append([m, n])
#                         else:
#                             print(k, '-->', v, "4")
#                             innerSec.append([k, v])
#                 else:
#                     print(key, '-->', value, 5)
#                     sec.append([key, str(value)])

#         # for indexed data, might wanna save it excel in a diff format instead of key value
#         elif sections in mylist and isinstance(asset[sections], list):
#             sec = wb.create_sheet(sections)
#             for item in asset[sections]:
#                 for k, v in item.items():
#                     print(k, '-->', v, "6")
#                     if isinstance(v, dict):
#                         innerSec = wb.create_sheet(sections + '_' + k)
#                         for x, y in v.items():
#                             print(x, '-->', y, "WOOO", "7")
#                             if isinstance(y, dict):
#                                 for m, n in y.items():
#                                     print(m, '-->', n, "8")
#                                     innerSec.append([m, n])
#                             else:
#                                 print(x, '-->', y, "9")
#                                 innerSec.append([x, y])
#                     else:
#                         print(k, '-->', v, "10")
#                         sec.append([k, v])
#         else:
#             print(sections, '-->', asset[sections], "11")
#             defaultSheet.append([sections, str(asset[sections])])
#     wb.save(filename="AssetObject.xlsx")

# # getAllDataFromAssetAndConverToExcelSheets()
